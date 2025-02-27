# python对接excel，读写操作
# 导包
import openpyxl
""""创建工作簿和写入数据"""
# 录入的数据（可以通过抓包获得，爬虫）
data = ["a", "b", "c", "d", "e"]
# 创建一个新的excel工作簿
test_excel = openpyxl.Workbook()  # 创建对象
# 在excel中创建工作表（测试表）
sheet = test_excel.create_sheet("测试")
# 向工作表中添加数据
for item in data:   # 一次添加一行
    sheet.append(data)
test_excel.save("测试.xlsx")
"""读取数据"""
# 打开工作簿
test_excel = openpyxl.load_workbook("测试.xlsx")
# 选择要操作的工作表
sheet = test_excel["测试"]
# 表格数据是二维列表，先遍历行，后遍历列
lst = []    # 存储行数据
for row in sheet.rows:
    sublst = [] # 存储单元格数据m,n
    for cell in row:    # cell是单元格
        sublst.append(cell.value)   # 将一行的单元格逐个放到sublst这个列表里面去
    lst.append(sublst)  # 每次存储一列的数据
for item in lst:    # 打印每行的数据
    print(item)


